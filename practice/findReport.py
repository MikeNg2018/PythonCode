#!python3
#功能：找出月报表中错误填写在CMS表中的报表类报障

import openpyxl
import re,os

os.chdir("C:\\work\\pythonpj\\")

#手动建立一个新Excel表，名为report.xlsx，用作储存提取数据，不存在则会报错
wbReport = openpyxl.load_workbook('report.xlsx')
newsheet = wbReport.get_active_sheet()
newsheet.title = 'Report'
newsheet = wbReport.get_sheet_by_name('Report')
#写入标题行
newsheet['A1'] = '影院编码'
newsheet['B1'] = '服务器ip'
newsheet['C1'] = '报障影院'
newsheet['D1'] = '业务系统分类'
newsheet['E1'] = '问题描述'
newsheet['F1'] = '解决方案'
newsheet['G1'] = '故障原因'
newsheet['H1'] = '交班人'
newsheet['I1'] = '处理时长'

#打开月报表，文件名为test.xlsx，表格
wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.get_sheet_by_name('CMS')
#月报表起始行
n = 2
#提取表起始行
m = 2
#遍历第五列的所有单元格，将含有“报表”的行复制到新表中
for f in sheet:
    check = sheet.cell(column=5, row=n)
    try:
        if re.match(r'\S*报表\S*',check.value):
            print(check.value)
            newsheet.cell(column=1, row=m).value = sheet.cell(column=1, row=n).value
            newsheet.cell(column=2, row=m).value = sheet.cell(column=2, row=n).value
            newsheet.cell(column=3, row=m).value = sheet.cell(column=3, row=n).value
            newsheet.cell(column=4, row=m).value = sheet.cell(column=4, row=n).value
            newsheet.cell(column=5, row=m).value = sheet.cell(column=5, row=n).value
            newsheet.cell(column=6, row=m).value = sheet.cell(column=6, row=n).value
            newsheet.cell(column=7, row=m).value = sheet.cell(column=7, row=n).value
            newsheet.cell(column=8, row=m).value = sheet.cell(column=8, row=n).value
            newsheet.cell(column=9, row=m).value = sheet.cell(column=9, row=n).value
            m = m + 1
        n = n + 1
    except TypeError:
        pass
wbReport.save('C:\\work\\pythonpj\\report.xlsx')
wbReport.close()
wb.close()